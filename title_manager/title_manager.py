"""
核心业务逻辑层 - 称号管理、Minecraft Team 交互
"""

import os
from typing import Optional

from mcdreforged.api.types import PluginServerInterface, CommandSource
from mcdreforged.api.rtext import RText, RColor, RStyle, RTextList, RAction

from .storage import Storage

# Minecraft 颜色名 -> RColor 映射
_COLOR_MAP = {
    'black': RColor.black, 'dark_blue': RColor.dark_blue,
    'dark_green': RColor.dark_green, 'dark_aqua': RColor.dark_aqua,
    'dark_red': RColor.dark_red, 'dark_purple': RColor.dark_purple,
    'gold': RColor.gold, 'gray': RColor.gray,
    'dark_gray': RColor.dark_gray, 'blue': RColor.blue,
    'green': RColor.green, 'aqua': RColor.aqua,
    'red': RColor.red, 'light_purple': RColor.light_purple,
    'yellow': RColor.yellow, 'white': RColor.white,
}

ADMIN_PERM = 2


def _resolve_color(color: str) -> RColor:
    """将颜色字符串解析为 RColor，支持名称和十六进制"""
    lower = color.lower()
    if lower in _COLOR_MAP:
        return _COLOR_MAP[lower]
    # 十六进制颜色: #RGB 或 #RRGGBB
    if lower.startswith('#'):
        try:
            return RColor.from_mc_value(lower)
        except (ValueError, TypeError):
            pass
    return RColor.white


def _format_title(name: str, color: str, bold: bool) -> RText:
    """创建带颜色和格式的称号 RText"""
    text = RText(name, _resolve_color(color))
    if bold:
        text.set_styles(RStyle.bold)
    return text


class TitleManager:
    """称号管理核心逻辑"""

    def __init__(self, server: PluginServerInterface, storage: Storage):
        self.server = server
        self.storage = storage

    # ---- Minecraft 命令执行 ----

    def _execute(self, cmd: str) -> None:
        self.server.execute(cmd)

    def _create_team(self, title_id: str, name: str, color: str, bold: str) -> None:
        """创建 Minecraft Team 并设置前缀"""
        bold_val = 'true' if bold.lower() == 'true' else 'false'
        self._execute(f'/team add {title_id}')
        self._execute(
            f'/team modify {title_id} prefix '
            f'{{"text":"{name} ","color":"{color}","bold":{bold_val}}}'
        )

    def _remove_team(self, title_id: str) -> None:
        self._execute(f'/team remove {title_id}')

    def _join_team(self, title_id: str, player: str) -> None:
        self._execute(f'/team join {title_id} {player}')

    def _leave_team(self, player: str) -> None:
        self._execute(f'/team leave {player}')

    # ===================== 玩家命令 =====================

    def cmd_help(self, source: CommandSource) -> None:
        """展示帮助信息"""
        source.reply(RText('======= 称号管理帮助 =======', RColor.gold))
        source.reply(RText('[ 玩家命令 ]', RColor.yellow))
        source.reply('  !!title help  - 显示帮助')
        source.reply('  !!title list [页码]  - 查看拥有的称号')
        source.reply('  !!title set <titleId>  - 佩戴称号')
        source.reply('  !!title leave  - 解除佩戴的称号')
        if source.has_permission(ADMIN_PERM):
            source.reply(RText('[ 管理命令 ]', RColor.yellow))
            source.reply('  !!title add <titleId> <名称> <颜色> <加粗>  - 创建称号(无需输入中括号)',)
            source.reply('  !!title remove <titleId>  - 删除称号')
            source.reply('  !!title join <玩家> <titleId>  - 配置并佩戴称号给玩家')
            source.reply('  !!title give <玩家> <titleId>  - 赠予玩家称号(不佩戴)')
            source.reply('  !!title delete <玩家> <titleId>  - 删除玩家称号')
            source.reply('  !!title show all  - 查看所有称号')
            source.reply('  !!title show player <玩家>  - 查看玩家称号')
            source.reply('  !!title show title <titleId>  - 查看称号拥有者')
            source.reply('  !!title move <旧玩家> <新玩家>  - 迁移玩家称号')
            source.reply('  !!title export  - 导出数据到Excel')
            source.reply('  !!title import  - 从Excel导入数据')

        # 快捷操作栏
        source.reply(RText('------- 快捷操作 -------', RColor.gold))
        source.reply(RTextList(
            RText('[查看称号]', RColor.green).h('查看你拥有的称号').c(RAction.suggest_command, '!!title list'),
            RText('  ', RColor.white),
            RText('[解除佩戴]', RColor.red).h('解除当前佩戴的称号').c(RAction.suggest_command, '!!title leave'),
        ))

    def cmd_list(self, source: CommandSource, page: int = 1) -> None:
        """展示玩家拥有的称号列表"""
        if not source.is_player:
            source.reply(RText('该命令仅限玩家使用', RColor.red))
            return

        player = source.player
        title_ids = self.storage.get_player_title_ids(player)
        wearing_id = self.storage.get_wearing(player)

        if not title_ids:
            source.reply(RText('你还没有任何称号', RColor.yellow))
            return

        per_page = 10
        total = len(title_ids)
        max_page = max(1, (total + per_page - 1) // per_page)
        page = max(1, min(page, max_page))
        start = (page - 1) * per_page
        end = min(start + per_page, total)

        source.reply(RText(f'=== 你的称号 ({page}/{max_page}) ===', RColor.gold))
        for idx, tid in enumerate(title_ids[start:end]):
            title = self.storage.get_title(tid)
            if title is None:
                continue
            display_num = start + idx + 1
            name_text = _format_title(title['name'], title['color'], title['bold'] == 'true')
            if tid == wearing_id:
                # 当前佩戴的称号：显示解除按钮
                wearing_tag = RText(' ★', RColor.green)
                action_btn = RText('[解除]', RColor.red).h('点击解除当前称号').c(RAction.suggest_command, '!!title leave')
            else:
                wearing_tag = RText('')
                action_btn = RText('[佩戴]', RColor.green).h('点击佩戴此称号').c(RAction.suggest_command, f'!!title set {tid}')
            # 称号名可点击佩戴
            name_text = name_text.h('点击佩戴此称号' if tid != wearing_id else '当前佩戴中').c(
                RAction.suggest_command, f'!!title set {tid}'
            )
            source.reply(RTextList(
                RText(f'  {display_num}. ', RColor.gray),
                RText(f'[{tid}] ', RColor.white),
                name_text,
                wearing_tag,
                RText(' ', RColor.white),
                action_btn,
            ))

        if page < max_page:
            source.reply(RText(f'输入 !!title list {page + 1} 查看下一页', RColor.gray))

    def cmd_set(self, source: CommandSource, title_id: str) -> None:
        """玩家佩戴某称号"""
        if not source.is_player:
            source.reply(RText('该命令仅限玩家使用', RColor.red))
            return

        player = source.player
        title = self.storage.get_title(title_id)
        if title is None:
            source.reply(RText(f'称号 [{title_id}] 不存在', RColor.red))
            return
        if not self.storage.player_has_title(player, title_id):
            source.reply(RText(f'你没有称号 [{title_id}] 的使用权', RColor.red))
            return

        wearing_id = self.storage.get_wearing(player)
        if wearing_id == title_id:
            source.reply(RText('你已佩戴该称号', RColor.yellow))
            return

        # 离开当前称号的 Team
        if wearing_id:
            self._leave_team(player)
        # 加入新称号的 Team
        self._join_team(title_id, player)
        self.storage.set_wearing(player, title_id)

        name_text = _format_title(title['name'], title['color'], title['bold'] == 'true')
        source.reply(RTextList(RText('已佩戴称号: ', RColor.green), name_text))

    def cmd_leave(self, source: CommandSource) -> None:
        """玩家解除当前佩戴的称号"""
        if not source.is_player:
            source.reply(RText('该命令仅限玩家使用', RColor.red))
            return

        player = source.player
        wearing_id = self.storage.get_wearing(player)
        if not wearing_id:
            source.reply(RText('你当前没有佩戴任何称号', RColor.yellow))
            return

        self._leave_team(player)
        self.storage.remove_wearing(player)
        title = self.storage.get_title(wearing_id)
        if title:
            name_text = _format_title(title['name'], title['color'], title['bold'] == 'true')
            source.reply(RTextList(RText('已解除称号: ', RColor.green), name_text))
        else:
            source.reply(RText(f'已解除称号: {wearing_id}', RColor.green))

    # ===================== 管理员命令 =====================

    def cmd_add(self, source: CommandSource, title_id: str, name: str, color: str, bold: str) -> None:
        """创建新称号"""
        if self.storage.get_title(title_id) is not None:
            source.reply(RText(f'称号ID [{title_id}] 已存在', RColor.red))
            return

        # 自动补齐中括号
        display_name = f'[{name}]'

        self.storage.add_title(title_id, display_name, color, bold)
        self._create_team(title_id, display_name, color, bold)
        name_text = _format_title(display_name, color, bold.lower() == 'true')
        source.reply(RTextList(RText(f'已创建称号 [{title_id}] ', RColor.green), name_text))

    def cmd_remove(self, source: CommandSource, title_id: str) -> None:
        """删除称号"""
        title = self.storage.get_title(title_id)
        if title is None:
            source.reply(RText(f'称号ID [{title_id}] 不存在', RColor.red))
            return

        # 移除所有玩家的该称号关联 & 解除佩戴
        removed_players = self.storage.remove_title_all_players(title_id)
        for p in removed_players:
            if self.storage.get_wearing(p) == title_id:
                self._leave_team(p)
                self.storage.remove_wearing(p)

        self.storage.remove_title(title_id)
        self._remove_team(title_id)
        name_text = _format_title(title['name'], title['color'], title['bold'] == 'true')
        source.reply(RTextList(RText(f'已删除称号 [{title_id}] ', RColor.green), name_text))

    def cmd_join(self, source: CommandSource, player_name: str, title_id: str) -> None:
        """手动配置并加入某称号给玩家（给予权限+佩戴）"""
        title = self.storage.get_title(title_id)
        if title is None:
            source.reply(RText(f'称号ID [{title_id}] 不存在', RColor.red))
            return

        # 确保玩家拥有该称号（忽略已拥有的情况）
        self.storage.add_player_title(player_name, title_id)

        # 先离开当前佩戴的称号
        wearing_id = self.storage.get_wearing(player_name)
        if wearing_id:
            self._leave_team(player_name)

        # 佩戴新称号
        self._join_team(title_id, player_name)
        self.storage.set_wearing(player_name, title_id)
        source.reply(RTextList(RText(f'已将称号 [{title_id}] ', RColor.green), _format_title(title['name'], title['color'], title['bold'] == 'true'), RText(f' 配置给玩家 {player_name}', RColor.green)))

    def cmd_give(self, source: CommandSource, player_name: str, title_id: str) -> None:
        """给某玩家新增一个称号但不佩戴"""
        title = self.storage.get_title(title_id)
        if title is None:
            source.reply(RText(f'称号ID [{title_id}] 不存在', RColor.red))
            return

        if not self.storage.add_player_title(player_name, title_id):
            source.reply(RTextList(RText(f'玩家 {player_name} 已拥有称号 [{title_id}] ', RColor.yellow), _format_title(title['name'], title['color'], title['bold'] == 'true')))
            return
        source.reply(RTextList(RText(f'已将称号 [{title_id}] ', RColor.green), _format_title(title['name'], title['color'], title['bold'] == 'true'), RText(f' 赠予玩家 {player_name}（未佩戴）', RColor.green)))

    def cmd_delete(self, source: CommandSource, player_name: str, title_id: str) -> None:
        """删除某玩家的某称号权限"""
        if not self.storage.player_has_title(player_name, title_id):
            source.reply(RText(f'玩家 {player_name} 没有称号 [{title_id}]', RColor.red))
            return

        # 如果玩家正在佩戴该称号，先解除
        if self.storage.get_wearing(player_name) == title_id:
            self._leave_team(player_name)
            self.storage.remove_wearing(player_name)

        self.storage.remove_player_title(player_name, title_id)
        title = self.storage.get_title(title_id)
        name_text = _format_title(title['name'], title['color'], title['bold'] == 'true') if title else RText(title_id)
        source.reply(RTextList(RText(f'已删除玩家 {player_name} 的称号 [{title_id}] ', RColor.green), name_text))

    def cmd_show_all(self, source: CommandSource) -> None:
        """展示当前服所有称号"""
        titles = self.storage.get_all_titles()
        if not titles:
            source.reply(RText('当前没有任何称号', RColor.yellow))
            return

        source.reply(RText('=== 所有称号 ===', RColor.gold))
        for t in titles:
            count = len(self.storage.get_title_players(t['id']))
            name_text = _format_title(t['name'], t['color'], t['bold'] == 'true')
            # 称号名可点击查看拥有者
            name_text = name_text.h(f'点击查看拥有称号 [{t["id"]}] 的玩家').c(
                RAction.suggest_command, f'!!title show title {t["id"]}'
            )
            source.reply(RTextList(
                RText(f'  [{t["id"]}] ', RColor.white),
                name_text,
                RText(f'  颜色:{t["color"]} 加粗:{t["bold"]} ({count}人)', RColor.gray),
            ))

    def cmd_show_player(self, source: CommandSource, player_name: str) -> None:
        """展示某玩家的所有称号"""
        title_ids = self.storage.get_player_title_ids(player_name)
        wearing_id = self.storage.get_wearing(player_name)

        if not title_ids:
            source.reply(RText(f'玩家 {player_name} 没有任何称号', RColor.yellow))
            return

        source.reply(RText(f'=== {player_name} 的称号 ===', RColor.gold))
        for tid in title_ids:
            title = self.storage.get_title(tid)
            if title is None:
                continue
            name_text = _format_title(title['name'], title['color'], title['bold'] == 'true')
            wearing_tag = RText(' ★', RColor.green) if tid == wearing_id else RText('')
            source.reply(RTextList(
                RText(f'  [{tid}] ', RColor.white),
                name_text,
                wearing_tag,
            ))

    def cmd_show_title(self, source: CommandSource, title_id: str) -> None:
        """展示拥有某称号的所有玩家"""
        title = self.storage.get_title(title_id)
        if title is None:
            source.reply(RText(f'称号ID [{title_id}] 不存在', RColor.red))
            return

        players = self.storage.get_title_players(title_id)
        name_text = _format_title(title['name'], title['color'], title['bold'] == 'true')
        source.reply(RTextList(
            RText('=== 称号 ', RColor.gold),
            RText(f'[{title_id}] ', RColor.white),
            name_text,
            RText(' 的拥有者 ===', RColor.gold),
        ))
        if not players:
            source.reply(RText('  暂无玩家拥有此称号', RColor.gray))
        else:
            for p in players:
                wearing_tag = ' ★' if self.storage.get_wearing(p) == title_id else ''
                source.reply(RText(f'  {p}{wearing_tag}', RColor.white))

    def cmd_move(self, source: CommandSource, old_name: str, new_name: str) -> None:
        """将改名前玩家的称号迁移到改名后玩家账户下"""
        wearing_id = self.storage.get_wearing(old_name)
        count = self.storage.move_player(old_name, new_name)
        if count == 0:
            source.reply(RText(f'玩家 {old_name} 没有任何称号数据', RColor.yellow))
            return

        # 更新 Minecraft Team 成员
        if wearing_id:
            self._leave_team(old_name)
            self._join_team(wearing_id, new_name)

        source.reply(RTextList(RText(f'已将 {old_name} 的 {count} 个称号迁移至 {new_name}', RColor.green)))

    # ===================== 数据导入导出 =====================

    def cmd_export(self, source: CommandSource) -> None:
        """导出称号数据到 Excel 文件"""
        try:
            from openpyxl import Workbook
        except ImportError:
            source.reply(RText('缺少 openpyxl 库，请在 Python 环境中执行 pip install openpyxl', RColor.red))
            return

        wb = Workbook()

        # ---- Sheet 1: 称号 ----
        ws_title = wb.active
        ws_title.title = '称号'
        ws_title.append(['称号ID', '名称', '颜色', '加粗'])
        for t in self.storage.get_all_titles():
            ws_title.append([t['id'], t['name'], t['color'], t['bold']])

        # ---- Sheet 2: 玩家称号 ----
        ws_player = wb.create_sheet('玩家称号')
        ws_player.append(['玩家名', '称号ID'])
        for p in self.storage.players:
            ws_player.append([p['playerName'], p['titleId']])

        # ---- Sheet 3: 佩戴状态 ----
        ws_wearing = wb.create_sheet('佩戴状态')
        ws_wearing.append(['玩家名', '佩戴称号ID'])
        for player_name, title_id in self.storage.wearing.items():
            ws_wearing.append([player_name, title_id])

        # 保存文件
        export_dir = self.storage.data_dir
        export_path = os.path.join(export_dir, 'title_data_export.xlsx')
        wb.save(export_path)

        source.reply(RTextList(
            RText('已导出称号数据到: ', RColor.green),
            RText(export_path, RColor.yellow),
        ))
        source.reply(RText(f'  称号: {len(self.storage.titles)} 条  玩家称号: {len(self.storage.players)} 条  佩戴: {len(self.storage.wearing)} 条', RColor.gray))

    def cmd_import(self, source: CommandSource) -> None:
        """从 Excel 文件导入称号数据"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            source.reply(RText('缺少 openpyxl 库，请在 Python 环境中执行 pip install openpyxl', RColor.red))
            return

        import_path = os.path.join(self.storage.data_dir, 'title_data_export.xlsx')
        if not os.path.exists(import_path):
            source.reply(RText(f'未找到导入文件: {import_path}', RColor.red))
            source.reply(RText('请将 Excel 文件放置到插件数据目录下，命名为 title_data_export.xlsx', RColor.yellow))
            return

        try:
            wb = load_workbook(import_path)
        except Exception as e:
            source.reply(RText(f'读取 Excel 文件失败: {e}', RColor.red))
            return

        # 读取数据
        titles_data = []
        players_data = []
        wearing_data = {}

        # Sheet 1: 称号
        if '称号' in wb.sheetnames:
            ws = wb['称号']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    titles_data.append({
                        'id': str(row[0]),
                        'name': str(row[1]) if row[1] else '',
                        'color': str(row[2]) if row[2] else 'white',
                        'bold': str(row[3]) if row[3] else 'false',
                    })

        # Sheet 2: 玩家称号
        if '玩家称号' in wb.sheetnames:
            ws = wb['玩家称号']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    players_data.append({
                        'playerName': str(row[0]),
                        'titleId': str(row[1]) if row[1] else '',
                    })

        # Sheet 3: 佩戴状态
        if '佩戴状态' in wb.sheetnames:
            ws = wb['佩戴状态']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and row[1] is not None:
                    wearing_data[str(row[0])] = str(row[1])

        # 确认覆盖
        source.reply(RText('即将导入以下数据（将覆盖当前所有数据）:', RColor.gold))
        source.reply(RText(f'  称号: {len(titles_data)} 条  玩家称号: {len(players_data)} 条  佩戴: {len(wearing_data)} 条', RColor.gray))
        source.reply(RTextList(
            RText('输入 ', RColor.yellow),
            RText('!!title import confirm', RColor.green).h('点击确认导入').c(RAction.suggest_command, '!!title import confirm'),
            RText(' 确认导入', RColor.yellow),
        ))

        # 暂存数据等待确认
        self._pending_import = {
            'titles': titles_data,
            'players': players_data,
            'wearing': wearing_data,
        }

    def cmd_import_confirm(self, source: CommandSource) -> None:
        """确认导入称号数据"""
        if not hasattr(self, '_pending_import') or self._pending_import is None:
            source.reply(RText('没有待导入的数据，请先执行 !!title import', RColor.red))
            return

        data = self._pending_import
        self._pending_import = None

        # 清除旧数据 - 先让所有在线玩家离开当前 Team
        for player_name, title_id in list(self.storage.wearing.items()):
            self._leave_team(player_name)

        # 移除所有旧 Team
        for t in self.storage.get_all_titles():
            self._remove_team(t['id'])

        # 覆盖数据
        self.storage.titles = data['titles']
        self.storage._save(self.storage.title_file, self.storage.titles)

        self.storage.players = data['players']
        self.storage._save(self.storage.player_file, self.storage.players)

        self.storage.wearing = data['wearing']
        self.storage._save(self.storage.wearing_file, self.storage.wearing)

        # 重新创建 Team
        for t in self.storage.get_all_titles():
            self._create_team(t['id'], t['name'], t['color'], t['bold'])

        # 重新让玩家加入 Team
        for player_name, title_id in self.storage.wearing.items():
            if self.storage.get_title(title_id):
                self._join_team(title_id, player_name)

        source.reply(RText('导入完成! 所有称号数据已从 Excel 覆盖', RColor.green))
        source.reply(RText(f'  称号: {len(self.storage.titles)} 条  玩家称号: {len(self.storage.players)} 条  佩戴: {len(self.storage.wearing)} 条', RColor.gray))

    # ===================== 事件处理 =====================

    def on_player_join(self, player: str) -> None:
        """玩家加入服务器时，确保佩戴状态与 Team 一致"""
        wearing_id = self.storage.get_wearing(player)
        if wearing_id and self.storage.get_title(wearing_id):
            self._join_team(wearing_id, player)
